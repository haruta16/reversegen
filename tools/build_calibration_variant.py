#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import random
import statistics
from collections import Counter, defaultdict, deque
from pathlib import Path
from typing import Iterable

import openpyxl


GRADES = range(6)
IGNORED_STAGE_EDITORS = {(1, "100001"), (2, "100002")}


def csv_escape(value: object) -> str:
    text = "" if value is None else str(value)
    if any(ch in text for ch in [",", "\n", "\r", '"']):
        return '"' + text.replace('"', '""') + '"'
    return text


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def is_valid_board(row: dict[str, str]) -> bool:
    if row.get("isMaxGradeProbe", "").strip() == "1":
        return False
    if not row.get("CompletionStatus", "Success").startswith("Success"):
        return False
    try:
        grade = int(str(row.get("grade", "")).strip())
    except ValueError:
        return False
    return grade in GRADES


def selected_rows(
    rows: Iterable[dict[str, str]],
    cap: int,
    min_per_level_grade: int = 0,
) -> tuple[list[dict[str, str]], Counter[tuple[str, int]]]:
    valid = [row for row in rows if is_valid_board(row)]
    available: Counter[tuple[str, int]] = Counter(
        (row.get("levelResId", "").strip(), int(row["grade"]))
        for row in valid
        if row.get("levelResId", "").strip()
    )
    counts: Counter[tuple[str, int]] = Counter()
    out: list[dict[str, str]] = []
    for row in valid:
        level = row.get("levelResId", "").strip()
        grade = int(row["grade"])
        key = (level, grade)
        if not level or available[key] < min_per_level_grade:
            continue
        if cap > 0 and counts[key] >= cap:
            continue
        counts[key] += 1
        out.append(row)
    return out, available


def sort_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    def key(row: dict[str, str]) -> tuple[int, int, int]:
        def to_int(value: str, default: int) -> int:
            try:
                return int(float(value))
            except (TypeError, ValueError):
                return default

        return (
            to_int(row.get("levelResId", ""), 0),
            to_int(row.get("grade", ""), 99),
            to_int(row.get("attemptIndex", ""), 0),
        )

    return sorted(rows, key=key)


def coverage_from_rows(rows: Iterable[dict[str, str]]) -> dict[str, Counter[int]]:
    coverage: dict[str, Counter[int]] = defaultdict(Counter)
    for row in rows:
        if not is_valid_board(row):
            continue
        level = row.get("levelResId", "").strip()
        if not level:
            continue
        coverage[level][int(row["grade"])] += 1
    return coverage


def grade_range(counter: Counter[int]) -> tuple[str | None, int | None, int | None, int]:
    grades = [g for g in GRADES if counter.get(g, 0) > 0]
    total = sum(counter.get(g, 0) for g in GRADES)
    if not grades:
        return None, None, None, total
    return ",".join(str(g) for g in grades), min(grades), max(grades), total


def clear_sheet_data(ws, start_row: int = 2) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_coverage_cells(ws, row: int, counter: Counter[int], note_override: str | None = None) -> None:
    ranges, low, high, total = grade_range(counter)
    ws.cell(row, 4).value = ranges
    ws.cell(row, 5).value = low
    ws.cell(row, 6).value = high
    ws.cell(row, 7).value = total
    for col, grade in enumerate(GRADES, start=8):
        ws.cell(row, col).value = counter.get(grade, 0)
    ws.cell(row, 14).value = note_override if note_override is not None else ("具备" if total else "无当前覆盖数据")


def refresh_distribution(wb, coverage: dict[str, Counter[int]]) -> None:
    ws = wb["地形G0-G5分布"]
    clear_sheet_data(ws, 2)
    for row, level in enumerate(sorted(coverage, key=lambda x: int(x)), start=2):
        counter = coverage[level]
        ranges, low, high, total = grade_range(counter)
        values = [level, total, low, high, ranges, *[counter.get(g, 0) for g in GRADES]]
        for col, value in enumerate(values, start=1):
            ws.cell(row, col).value = value


def refresh_summary(wb, coverage: dict[str, Counter[int]], input_rows: int, valid_rows: int) -> None:
    ws = wb["难度汇总"]
    clear_sheet_data(ws, 2)
    ordered_levels = sorted(coverage, key=lambda x: int(x))
    for row, grade in enumerate(GRADES, start=2):
        levels = [level for level in ordered_levels if coverage[level].get(grade, 0) > 0]
        ws.cell(row, 1).value = f"G{grade}"
        ws.cell(row, 2).value = len(levels)
        ws.cell(row, 3).value = sum(coverage[level].get(grade, 0) for level in levels)
        ws.cell(row, 4).value = ",".join(levels)

    if "统计说明" not in wb.sheetnames:
        return
    info = wb["统计说明"]
    values = {
        "输入行数": input_rows,
        "有效牌局行数": valid_rows,
        "过滤探测行": None,
        "过滤失败行": None,
        "过滤无效难度行": None,
        "地形数": len(coverage),
    }
    for row in range(2, info.max_row + 1):
        key = info.cell(row, 1).value
        if key in values:
            info.cell(row, 2).value = values[key]


def refresh_mainline(wb, coverage: dict[str, Counter[int]]) -> None:
    ws = wb["主线关难度覆盖"]
    for row in range(2, ws.max_row + 1):
        stage = ws.cell(row, 1).value
        editor = str(ws.cell(row, 3).value or "").strip()
        counter = coverage.get(editor, Counter())
        note = "无需处理" if (stage, editor) in IGNORED_STAGE_EDITORS else None
        write_coverage_cells(ws, row, counter, note)


def refresh_endless(wb, coverage: dict[str, Counter[int]]) -> None:
    ws = wb["无尽关难度覆盖"]
    for row in range(2, ws.max_row + 1):
        editor = str(ws.cell(row, 2).value or "").strip()
        counter = coverage.get(editor, Counter())
        ranges, low, high, total = grade_range(counter)
        ws.cell(row, 3).value = ranges
        ws.cell(row, 4).value = low
        ws.cell(row, 5).value = high
        ws.cell(row, 6).value = total
        for col, grade in enumerate(GRADES, start=7):
            ws.cell(row, col).value = counter.get(grade, 0)
        ws.cell(row, 13).value = "具备" if total else "无当前覆盖数据"


def parse_sequence_grades(value: object) -> list[int]:
    text = "" if value is None else str(value).strip()
    if not text:
        return []
    grades: list[int] = []
    for part in text.replace("，", ",").split(","):
        part = part.strip()
        if not part:
            continue
        try:
            grade = int(float(part))
        except ValueError:
            continue
        if grade in GRADES:
            grades.append(grade)
    return grades


def refresh_front80(wb, coverage: dict[str, Counter[int]]) -> list[dict[str, object]]:
    ws = wb["前80关在线胜率"]
    issues: list[dict[str, object]] = []
    headers = {str(ws.cell(4, col).value): col for col in range(1, ws.max_column + 1)}
    editor_col = headers.get("EditorID", 6)
    sequence_col = headers.get("近似模拟GradeSequence", 4)
    required_col = headers.get("要求Grade", 7)
    note_col = headers.get("备注", 8)

    for row in range(5, ws.max_row + 1):
        stage = ws.cell(row, 1).value
        editor = str(ws.cell(row, editor_col).value or "").strip()
        if not editor:
            continue
        required = parse_sequence_grades(ws.cell(row, sequence_col).value)
        ws.cell(row, required_col).value = ",".join(f"G{g}" for g in required)
        counter = coverage.get(editor, Counter())
        missing = [g for g in required if counter.get(g, 0) <= 0]
        if (stage, editor) in IGNORED_STAGE_EDITORS:
            note = "无需处理"
            missing = []
        elif not counter:
            note = f"地形{editor}无当前覆盖数据"
            if missing:
                note += "；缺失" + ",".join(f"G{g}" for g in missing)
        elif missing:
            note = "缺失" + ",".join(f"G{g}" for g in missing)
        else:
            note = "具备"
        ws.cell(row, note_col).value = note
        if missing or (not counter and (stage, editor) not in IGNORED_STAGE_EDITORS):
            issues.append({"stage": stage, "editorId": editor, "sequence": ws.cell(row, sequence_col).value, "note": note})
    return issues


def read_zones(wb) -> list[dict[str, object]]:
    ws = wb["无尽难度集"]
    zones: list[dict[str, object]] = []
    for row in range(5, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name:
            continue
        sequences = [
            "" if ws.cell(row, col).value is None else str(ws.cell(row, col).value)
            for col in range(2, 7)
        ]
        if not any(seq for seq in sequences):
            continue
        first_grades = [(parse_sequence_grades(seq) or [None])[0] for seq in sequences]
        zones.append({"name": str(name), "sequences": sequences, "firstGrades": first_grades})
    return zones[:6]


def endless_pool_from_sheet(wb) -> list[dict[str, object]]:
    ws = wb["无尽关难度覆盖"]
    pool: list[dict[str, object]] = []
    for row in range(2, ws.max_row + 1):
        editor = str(ws.cell(row, 2).value or "").strip()
        if not editor:
            continue
        grades = [g for g in GRADES if (ws.cell(row, 7 + g).value or 0) > 0]
        pool.append({"editor": editor, "grades": grades, "range": ",".join(str(g) for g in grades)})
    return pool


def simulate_endless(wb, groups: int = 10000, cooldown: int = 10, seed: int = 20260625) -> dict[str, object]:
    zones = read_zones(wb)
    pool = endless_pool_from_sheet(wb)
    by_grade: dict[int, list[str]] = {g: [] for g in GRADES}
    for item in pool:
        for grade in item["grades"]:
            by_grade[grade].append(item["editor"])

    rng = random.Random(seed)
    zone_bag: list[int] = []
    cooldown_queue: deque[str] = deque()
    cooldown_set: set[str] = set()
    selected_counts: Counter[str] = Counter()
    zone_counts: Counter[str] = Counter()
    demand_counts: Counter[int] = Counter()
    cooldown_fallbacks = 0
    missing_grade = 0

    def next_zone_index() -> int:
        nonlocal zone_bag
        if not zone_bag:
            zone_bag = list(range(len(zones)))
            rng.shuffle(zone_bag)
        return zone_bag.pop()

    for _ in range(groups):
        zone_idx = next_zone_index()
        zone = zones[zone_idx]
        zone_counts[str(zone["name"])] += 1
        for grade in zone["firstGrades"]:
            if grade is None:
                continue
            demand_counts[grade] += 1
            candidates = by_grade.get(grade, [])
            if not candidates:
                missing_grade += 1
                continue
            available = [editor for editor in candidates if editor not in cooldown_set]
            if not available:
                available = candidates
                cooldown_fallbacks += 1
            editor = rng.choice(available)
            selected_counts[editor] += 1
            cooldown_queue.append(editor)
            cooldown_set.add(editor)
            while len(cooldown_queue) > cooldown:
                old = cooldown_queue.popleft()
                if old not in cooldown_queue:
                    cooldown_set.discard(old)

    counts = [selected_counts[item["editor"]] for item in pool]
    mean = statistics.mean(counts) if counts else 0
    variance = statistics.pvariance(counts) if counts else 0
    stdev = statistics.pstdev(counts) if counts else 0
    return {
        "groups": groups,
        "stages": groups * 5,
        "pool": pool,
        "zones": zones,
        "zoneCounts": zone_counts,
        "demandCounts": demand_counts,
        "selectedCounts": selected_counts,
        "cooldownFallbacks": cooldown_fallbacks,
        "missingGrade": missing_grade,
        "mean": mean,
        "variance": variance,
        "stdev": stdev,
        "cv": stdev / mean if mean else 0,
        "max": max(counts) if counts else 0,
        "min": min(counts) if counts else 0,
        "byGradeCandidates": {g: len(by_grade[g]) for g in GRADES},
    }


def refresh_simulation(wb) -> dict[str, object]:
    result = simulate_endless(wb)
    ws = wb["无尽池模拟10000组"]
    clear_sheet_data(ws, 4)

    rows: list[list[object]] = [
        ["模拟难度集数", result["groups"], "模拟关卡数", result["stages"], "地形数", len(result["pool"]), "冷却关数", 10],
        ["平均选中次数", result["mean"], "标准差", result["stdev"], "方差", result["variance"], "变异系数", result["cv"]],
        ["最高选中次数", result["max"], "最低选中次数", result["min"], "冷却退回次数", result["cooldownFallbacks"], "缺可选难度次数", result["missingGrade"]],
        [],
        ["难度集首档", None, None, None, None, None, None, None, "难度需求/候选池", None, None, None, None, None, "偏高候选Top"],
        ["组", "第1关", "第2关", "第3关", "第4关", "第5关", "模拟抽中组数", None, "难度", "需求次数", "候选地形数", "理论平均每候选", None, None, "EditorID", "可用难度", "选中次数", "占比", "备注"],
    ]
    max_rows = max(len(result["zones"]), 6, len(result["pool"]))
    top = sorted(result["pool"], key=lambda item: result["selectedCounts"][item["editor"]], reverse=True)
    for i in range(max_rows):
        row: list[object] = []
        if i < len(result["zones"]):
            zone = result["zones"][i]
            row.extend([zone["name"], *zone["firstGrades"], result["zoneCounts"][str(zone["name"])]])
        else:
            row.extend([None] * 7)
        row.append(None)
        if i < 6:
            demand = result["demandCounts"][i]
            candidates = result["byGradeCandidates"][i]
            row.extend([f"G{i}", demand, candidates, demand / candidates if candidates else None])
        else:
            row.extend([None] * 4)
        row.extend([None, None])
        if i < min(20, len(top)):
            item = top[i]
            selected = result["selectedCounts"][item["editor"]]
            row.extend([item["editor"], item["range"], selected, selected / result["stages"] if result["stages"] else 0, ""])
        rows.append(row)

    for r, values in enumerate(rows, start=4):
        for c, value in enumerate(values, start=1):
            ws.cell(r, c).value = value
    return {
        "groups": result["groups"],
        "stages": result["stages"],
        "poolSize": len(result["pool"]),
        "mean": result["mean"],
        "stdev": result["stdev"],
        "missingGrade": result["missingGrade"],
        "cooldownFallbacks": result["cooldownFallbacks"],
        "byGradeCandidates": result["byGradeCandidates"],
    }


def build_variant(args: argparse.Namespace) -> None:
    headers, source_rows = read_csv(args.source_csv)
    selected, available = selected_rows(source_rows, args.cap, args.min_per_level_grade)
    selected = sort_rows(selected)
    write_csv(args.output_csv, headers, selected)

    coverage = coverage_from_rows(selected)
    wb = openpyxl.load_workbook(args.template_workbook)
    refresh_distribution(wb, coverage)
    refresh_summary(wb, coverage, input_rows=len(source_rows), valid_rows=len(selected))
    refresh_mainline(wb, coverage)
    refresh_endless(wb, coverage)
    front80_issues = refresh_front80(wb, coverage)
    simulation = refresh_simulation(wb)
    wb.save(args.output_workbook)

    report = {
        "variant": args.variant_name,
        "sourceCsv": str(args.source_csv),
        "outputCsv": str(args.output_csv),
        "templateWorkbook": str(args.template_workbook),
        "outputWorkbook": str(args.output_workbook),
        "capPerLevelGrade": args.cap,
        "minPerLevelGrade": args.min_per_level_grade,
        "sourceRows": len(source_rows),
        "outputRows": len(selected),
        "eligibleLevelGradePairs": sum(1 for count in available.values() if count >= args.min_per_level_grade),
        "excludedLevelGradePairs": sum(1 for count in available.values() if count < args.min_per_level_grade),
        "levels": len(coverage),
        "gradeRows": {f"G{g}": sum(c.get(g, 0) for c in coverage.values()) for g in GRADES},
        "gradeLevelCoverage": {f"G{g}": sum(1 for c in coverage.values() if c.get(g, 0) > 0) for g in GRADES},
        "front80Issues": front80_issues,
        "simulation": simulation,
    }
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a calibration workbook variant from a board CSV.")
    parser.add_argument("--source-csv", type=Path, required=True)
    parser.add_argument("--template-workbook", type=Path, default=Path("output/无尽关校准工具.xlsx"))
    parser.add_argument("--output-csv", type=Path, required=True)
    parser.add_argument("--output-workbook", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--variant-name", default="variant")
    parser.add_argument("--cap", type=int, default=10, help="每个地形/难度最多保留数量；0 表示不设上限")
    parser.add_argument("--min-per-level-grade", type=int, default=0, help="地形/难度达到该数量才整体保留")
    return parser.parse_args()


if __name__ == "__main__":
    build_variant(parse_args())
