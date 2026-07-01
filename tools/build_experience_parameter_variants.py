#!/usr/bin/env python3
"""Build checklist sheets that isolate one experience parameter at a time."""

from __future__ import annotations

import argparse
import itertools
import math
import os
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from build_experience_sample_workbook import (
    DEFAULT_ACCEPTED,
    DEFAULT_BACKFILL,
    DEFAULT_BASE,
    DEFAULT_METRICS,
    FEATURES,
    ROOT,
    SOURCE_EXCLUDED,
    build_pool,
    read_csv,
)
from fill_experience_checklist import DEFAULT_CHECKLIST, DEFAULT_FINAL_SELECTION


DEFAULT_OUTPUT = ROOT / "output/体验清单_单参数差异版本.xlsx"
VARIANTS = [
    ("花色系数差异", "color_ratio", 5),
    ("分布参数差异", "spread", 6),
    ("债务参数差异", "debt", 7),
    ("Optimal胜率差异", "optimal_win", 8),
    ("胜局断色差异", "starvation_win", 9),
    ("失败剩余差异", "remaining_loss", 10),
]


def normalized_vectors(rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, float | None]], dict[str, tuple[float, float]]]:
    ranges: dict[str, tuple[float, float]] = {}
    for feature in FEATURES:
        values = [row[feature] for row in rows if row.get(feature) is not None]
        ranges[feature] = (min(values), max(values)) if values else (0.0, 0.0)
    vectors: dict[str, dict[str, float | None]] = {}
    for row in rows:
        vector: dict[str, float | None] = {}
        for feature in FEATURES:
            value = row.get(feature)
            low, high = ranges[feature]
            vector[feature] = None if value is None else (0.5 if high == low else (value - low) / (high - low))
        vectors[row["ReplayCode"]] = vector
    return vectors, ranges


def other_distance(combo: tuple[dict[str, Any], ...], vectors: dict[str, dict[str, float | None]], target: str) -> float:
    distances = []
    for left, right in itertools.combinations(combo, 2):
        deltas = []
        for feature in FEATURES:
            if feature == target:
                continue
            a = vectors[left["ReplayCode"]][feature]
            b = vectors[right["ReplayCode"]][feature]
            if a is not None and b is not None:
                deltas.append((a - b) ** 2)
        distances.append(math.sqrt(sum(deltas) / len(deltas)) if deltas else 0.0)
    return sum(distances) / len(distances) if distances else 0.0


def select_target_variant(rows: list[dict[str, Any]], target: str, count: int) -> list[dict[str, Any]]:
    vectors, _ = normalized_vectors(rows)
    best_combo = None
    best_score = None
    for combo in itertools.combinations(rows, count):
        values = [vectors[row["ReplayCode"]][target] for row in combo]
        valid = sorted(value for value in values if value is not None)
        target_range = valid[-1] - valid[0] if len(valid) >= 2 else 0.0
        minimum_gap = min((right - left for left, right in zip(valid, valid[1:])), default=0.0)
        similarity_penalty = other_distance(combo, vectors, target)
        sources = {row["source"] for row in combo}
        score = (
            len(valid),
            round(target_range, 12),
            round(minimum_gap, 12),
            -round(similarity_penalty, 12),
            int(SOURCE_EXCLUDED in sources),
            len(sources),
        )
        if best_score is None or score > best_score:
            best_combo, best_score = combo, score
    if best_combo is None:
        raise ValueError(f"候选不足，无法选择 {count} 局")
    return sorted(best_combo, key=lambda row: (row.get(target) is None, row.get(target) if row.get(target) is not None else math.inf, row["ReplayCode"]))


def reset_sheet_values(worksheet) -> None:
    for row in range(3, 57):
        for column in range(2, 11):
            cell = worksheet.cell(row, column)
            cell.value = None
            cell.comment = None


def write_variant_sheet(worksheet, selected: list[dict[str, Any]], target: str, target_column: int, final_codes: set[str]) -> tuple[int, list[str]]:
    reset_sheet_values(worksheet)
    limitations: list[str] = []
    final_count = 0
    cursor = 3
    for level in ["100014", "100088", "100093"]:
        for grade in range(6):
            rows = [row for row in selected if row["level"] == level and row["grade_num"] == grade]
            target_values = [row[target] for row in rows if row.get(target) is not None]
            if len(target_values) < 3:
                limitations.append(f"{level}/G{grade}: 有效值 {len(target_values)}/3")
            elif max(target_values) == min(target_values):
                limitations.append(f"{level}/G{grade}: 目标参数无跨度")
            for row in rows:
                is_final = row["ReplayCode"] in final_codes
                values = [
                    row["ReplayCode"],
                    f"✓ G{grade}" if is_final else f"G{grade}",
                    row["color_count"], row["color_ratio"], row["spread"], row["debt"], row["optimal_win"],
                    row["starvation_win"], row["remaining_loss"],
                ]
                for column, value in enumerate(values, 2):
                    worksheet.cell(cursor, column, value)
                target_value = row.get(target)
                target_text = "空白（该机器人结果不适用）" if target_value is None else f"{target_value:.4f}"
                worksheet.cell(cursor, 2).comment = Comment(
                    f"本页最大距离参数：{worksheet.title}\n目标参数值：{target_text}\n"
                    f"样本来源：{row['source']}\n正式资源：{'✓ 已入选' if is_final else '未入选（体验对照）'}",
                    "Codex",
                )
                if is_final:
                    worksheet.cell(cursor, 3).fill = PatternFill("solid", fgColor="2E7D32")
                    worksheet.cell(cursor, 3).font = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")
                    final_count += 1
                else:
                    worksheet.cell(cursor, 3).fill = copy(worksheet.cell(3, 4).fill)
                worksheet.cell(cursor, 5).number_format = "0.000"
                worksheet.cell(cursor, 6).number_format = "0.000"
                worksheet.cell(cursor, 7).number_format = "0.000"
                worksheet.cell(cursor, 8).number_format = "0.0%"
                worksheet.cell(cursor, 9).number_format = "0.00"
                worksheet.cell(cursor, 10).number_format = "0.0%"
                worksheet.cell(cursor, target_column).fill = PatternFill("solid", fgColor="FFF2CC")
                cursor += 1
    worksheet.cell(2, target_column).fill = PatternFill("solid", fgColor="C78B32")
    worksheet.cell(2, target_column).font = Font(name="PingFang SC", bold=True, color="FFFFFF")
    worksheet.cell(1, 4).comment = Comment(
        f"本页以“{worksheet.cell(2, target_column).value}”最大距离为首要目标，其他体验参数尽量相似。\n"
        + ("无法形成完整跨度的组：\n" + "\n".join(limitations) if limitations else "所有地形档位均具有有效跨度。"),
        "Codex",
    )
    return final_count, limitations


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=DEFAULT_CHECKLIST)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--per-grade", type=int, default=3)
    args = parser.parse_args()

    levels = ["100014", "100088", "100093"]
    pool = build_pool(DEFAULT_BASE, DEFAULT_METRICS, DEFAULT_ACCEPTED, DEFAULT_BACKFILL)
    groups = defaultdict(list)
    for row in pool:
        if row["level"] in levels and 0 <= row["grade_num"] <= 5:
            groups[(row["level"], row["grade_num"])].append(row)
    final_codes = {row["ReplayCode"] for row in read_csv(DEFAULT_FINAL_SELECTION)}

    workbook = load_workbook(args.template)
    template_sheet = workbook[workbook.sheetnames[0]]
    for extra in workbook.worksheets[1:]:
        workbook.remove(extra)
    variant_sheets = [template_sheet] + [workbook.copy_worksheet(template_sheet) for _ in VARIANTS[1:]]
    results = {}
    for index, ((sheet_name, feature, column), worksheet) in enumerate(zip(VARIANTS, variant_sheets)):
        worksheet.title = sheet_name
        selected = []
        for level in levels:
            for grade in range(6):
                selected.extend(select_target_variant(groups[(level, grade)], feature, args.per_grade))
        final_count, limitations = write_variant_sheet(worksheet, selected, feature, column, final_codes)
        worksheet.sheet_properties.tabColor = ["4F81BD", "70AD47", "A5A5A5", "5B9BD5", "ED7D31", "C55A11"][index]
        results[sheet_name] = {"rows": len(selected), "final_selected": final_count, "limited_groups": len(limitations)}

    args.output.parent.mkdir(parents=True, exist_ok=True)
    temporary = args.output.with_name(f".{args.output.name}.tmp.xlsx")
    workbook.save(temporary)
    os.replace(temporary, args.output)
    print({"output": str(args.output), "variants": results})


if __name__ == "__main__":
    main()
