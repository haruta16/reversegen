#!/usr/bin/env python3
"""Build a diverse, playable sample workbook across terrain and grade."""

from __future__ import annotations

import argparse
import csv
import itertools
import math
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BASE = ROOT / "output/100003～100071_100073+_合并去少_含补缺_每档最多10_G5替换.csv"
DEFAULT_METRICS = ROOT / "output/replay导出_G5替换/selection_optimal.csv"
DEFAULT_ACCEPTED = ROOT / "output/replay导出_G5替换/selection_Optimal体验筛选_v1.csv"
DEFAULT_BACKFILL = ROOT / "output/generation_feature/runs/optimal_experience_backfill_20260629/01_generation/backfill.csv"
DEFAULT_OUTPUT = ROOT / "output/体验牌局抽样_三地形全难度.xlsx"

SOURCE_KEPT = "原始保留"
SOURCE_EXCLUDED = "原始-Optimal排除"
SOURCE_BACKFILL = "Optimal补档命中"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def number(row: dict[str, Any], *keys: str) -> float | None:
    for key in keys:
        value = row.get(key)
        if value in (None, ""):
            continue
        try:
            parsed = float(value)
            return parsed if math.isfinite(parsed) else None
        except (TypeError, ValueError):
            continue
    return None


def normalized_rate(row: dict[str, Any], ratio_key: str, percent_key: str) -> float | None:
    ratio = number(row, ratio_key)
    if ratio is not None:
        return ratio / 100 if ratio > 1 else ratio
    percent = number(row, percent_key)
    return None if percent is None else percent / 100


def enrich(row: dict[str, Any], source: str) -> dict[str, Any]:
    item = dict(row)
    item["source"] = source
    item["level"] = str(item.get("levelResId", ""))
    item["grade_num"] = int(float(item.get("grade", 0)))
    item["tile_count"] = int(number(item, "totalTiles", "地形总牌数") or 0)
    item["color_count"] = int(number(item, "colorCount", "ElementCount") or 0)
    triplet_capacity = item["tile_count"] // 3
    item["color_ratio"] = item["color_count"] / triplet_capacity if triplet_capacity else None
    item["spread"] = number(item, "spreadParam")
    item["debt"] = number(item, "debtPersistenceWeight")
    item["optimal_win"] = normalized_rate(item, "optimalWinRate", "最优机器人胜率(%)")
    item["starvation_win"] = number(item, "optimalStarvationOnWin", "最优机器人胜局平均断色次数")
    losses = number(item, "optimalLosses", "最优机器人负局数")
    item["remaining_loss"] = None if not losses else normalized_rate(
        item, "optimalRemainingRatioOnLoss", "最优机器人负局平均剩余牌比例(%)"
    )
    return item


def build_pool(base_path: Path, metrics_path: Path, accepted_path: Path, backfill_path: Path) -> list[dict[str, Any]]:
    metrics = {row["ReplayCode"]: row for row in read_csv(metrics_path)}
    accepted = {row["ReplayCode"] for row in read_csv(accepted_path)}
    pool: list[dict[str, Any]] = []
    for base_row in read_csv(base_path):
        merged = dict(base_row)
        merged.update({key: value for key, value in metrics.get(base_row["ReplayCode"], {}).items() if key not in merged or not merged[key]})
        source = SOURCE_KEPT if base_row["ReplayCode"] in accepted else SOURCE_EXCLUDED
        pool.append(enrich(merged, source))
    pool.extend(enrich(row, SOURCE_BACKFILL) for row in read_csv(backfill_path))
    unique: dict[str, dict[str, Any]] = {}
    for row in pool:
        code = row.get("ReplayCode", "")
        if code and code not in unique:
            unique[code] = row
    return list(unique.values())


FEATURES = ["color_ratio", "spread", "debt", "optimal_win", "starvation_win", "remaining_loss"]


def feature_vectors(rows: list[dict[str, Any]]) -> dict[str, list[float]]:
    ranges: dict[str, tuple[float, float]] = {}
    for feature in FEATURES:
        values = [row[feature] for row in rows if row.get(feature) is not None]
        ranges[feature] = (min(values), max(values)) if values else (0.0, 0.0)
    vectors: dict[str, list[float]] = {}
    for row in rows:
        vector: list[float] = []
        for feature in FEATURES:
            low, high = ranges[feature]
            value = row.get(feature)
            vector.append(0.5 if value is None or high == low else (value - low) / (high - low))
        vectors[row["ReplayCode"]] = vector
    return vectors


def distance(a: list[float], b: list[float]) -> float:
    return math.sqrt(sum((left - right) ** 2 for left, right in zip(a, b)))


def select_diverse(rows: list[dict[str, Any]], count: int) -> list[dict[str, Any]]:
    if len(rows) < count:
        raise ValueError(f"候选只有 {len(rows)} 条，无法选择 {count} 条")
    vectors = feature_vectors(rows)
    best_combo: tuple[dict[str, Any], ...] | None = None
    best_score = -1.0
    for combo in itertools.combinations(rows, count):
        sources = {row["source"] for row in combo}
        source_score = len(sources) * 8
        if SOURCE_EXCLUDED in sources:
            source_score += 2
        if SOURCE_KEPT in sources:
            source_score += 1
        pair_distances = [
            distance(vectors[left["ReplayCode"]], vectors[right["ReplayCode"]])
            for left, right in itertools.combinations(combo, 2)
        ]
        completeness = sum(sum(row.get(feature) is not None for feature in FEATURES) for row in combo) / (count * len(FEATURES))
        score = source_score + (sum(pair_distances) / len(pair_distances) if pair_distances else 0) + completeness
        if score > best_score:
            best_combo, best_score = combo, score
    assert best_combo is not None
    source_order = {SOURCE_KEPT: 0, SOURCE_EXCLUDED: 1, SOURCE_BACKFILL: 2}
    return sorted(best_combo, key=lambda row: (source_order[row["source"]], row["ReplayCode"]))


FEATURE_LABELS = {
    "color_ratio": "花色系数",
    "spread": "分布",
    "debt": "债务",
    "optimal_win": "Optimal胜率",
    "starvation_win": "断色",
    "remaining_loss": "失败剩余",
}


def difference_label(row: dict[str, Any], candidates: list[dict[str, Any]]) -> str:
    ranked: list[tuple[float, str]] = []
    for feature in FEATURES:
        values = [item[feature] for item in candidates if item.get(feature) is not None]
        value = row.get(feature)
        if value is None or len(values) < 2 or max(values) == min(values):
            continue
        center = median(values)
        deviation = (value - center) / (max(values) - min(values))
        direction = "高" if deviation >= 0 else "低"
        ranked.append((abs(deviation), direction + FEATURE_LABELS[feature]))
    ranked.sort(reverse=True)
    return " / ".join(label for _, label in ranked[:2]) or "同参数对照"


def add_title(ws, title: str, columns: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = ws.cell(1, 1, title)
    cell.font = Font(name="PingFang SC", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="243447")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30


def style_table(ws, header_row: int, widths: list[float]) -> None:
    thin = Side(style="thin", color="CBD2D9")
    for cell in ws[header_row]:
        cell.font = Font(name="PingFang SC", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3F5F7A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.font = Font(name="PingFang SC", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(widths))}{ws.max_row}"


def build_workbook(pool: list[dict[str, Any]], levels: list[str], per_grade: int, output: Path, sources: list[Path]) -> dict[str, Any]:
    groups: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in pool:
        if row["level"] in levels and 0 <= row["grade_num"] <= 5:
            groups[(row["level"], row["grade_num"])].append(row)
    missing = [(level, grade, len(groups[(level, grade)])) for level in levels for grade in range(6) if len(groups[(level, grade)]) < per_grade]
    if missing:
        raise ValueError(f"以下地形档位候选不足: {missing}")

    selected: list[dict[str, Any]] = []
    for level in levels:
        for grade in range(6):
            candidates = groups[(level, grade)]
            for row in select_diverse(candidates, per_grade):
                picked = dict(row)
                picked["difference"] = difference_label(row, candidates)
                selected.append(picked)

    wb = Workbook()
    ws = wb.active
    ws.title = "体验样本"
    headers = [
        "地形ID", "地形牌数", "ReplayCode", "实际难度分档", "花色数", "实际花色系数",
        "花色分布参数", "债务持续参数", "短视最优策略胜率", "胜局思考量（断色数）",
        "失败剩余率", "样本来源 / Optimal结果", "差异特征",
    ]
    add_title(ws, "三地形 · G0-G5 差异体验样本", len(headers))
    ws.cell(2, 1, f"每个地形每个难度选择 {per_grade} 局；来源包含原始保留、Optimal 排除与 Optimal 补档。")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1).font = Font(name="PingFang SC", size=10, color="57606A")
    ws.append(headers)
    fills = {"100014": "F4F9FF", "100088": "F5FBF6", "100093": "FFF8F0"}
    source_fonts = {SOURCE_KEPT: "1A7F37", SOURCE_EXCLUDED: "CF222E", SOURCE_BACKFILL: "8250DF"}
    for row in selected:
        ws.append([
            row["level"], row["tile_count"], row["ReplayCode"], f"G{row['grade_num']}", row["color_count"], row["color_ratio"],
            row["spread"], row["debt"], row["optimal_win"], row["starvation_win"], row["remaining_loss"], row["source"], row["difference"],
        ])
        current = ws.max_row
        for cell in ws[current]:
            cell.fill = PatternFill("solid", fgColor=fills.get(row["level"], "FFFFFF"))
        ws.cell(current, 12).font = Font(name="PingFang SC", size=10, bold=True, color=source_fonts[row["source"]])
        for col in (6, 7, 8):
            ws.cell(current, col).number_format = "0.000"
        for col in (9, 11):
            ws.cell(current, col).number_format = "0.0%"
        ws.cell(current, 10).number_format = "0.00"
        ws.row_dimensions[current].height = 32
    style_table(ws, 3, [12, 11, 52, 13, 10, 14, 14, 14, 17, 20, 14, 22, 24])

    summary = wb.create_sheet("选择说明")
    add_title(summary, "样本选择说明", 8)
    notes = [
        "目标：为三个不同 tile 规模的地形提供完整 G0-G5 体验样本，每档三局。",
        "优先级：先最大化样本来源差异，再最大化花色系数、分布、债务、Optimal胜率、断色和失败剩余率的参数距离。",
        "失败剩余率仅在 Optimal 存在失败局时填写；全胜牌局留空，避免把无失败误写成 100%。",
        "实际花色系数 = 花色数 ÷ floor(地形牌数 / 3)。",
    ]
    for index, note in enumerate(notes, 3):
        summary.cell(index, 1, note)
        summary.merge_cells(start_row=index, start_column=1, end_row=index, end_column=8)
    row_index = 8
    summary.cell(row_index, 1, "地形ID")
    summary.cell(row_index, 2, "地形牌数")
    summary.cell(row_index, 3, "候选总数")
    summary.cell(row_index, 4, SOURCE_KEPT)
    summary.cell(row_index, 5, SOURCE_EXCLUDED)
    summary.cell(row_index, 6, SOURCE_BACKFILL)
    summary.cell(row_index, 7, "覆盖档位")
    summary.cell(row_index, 8, "入表数量")
    for level in levels:
        terrain_rows = [row for row in pool if row["level"] == level]
        counts = Counter(row["source"] for row in terrain_rows)
        summary.append([
            level, terrain_rows[0]["tile_count"], len(terrain_rows), counts[SOURCE_KEPT], counts[SOURCE_EXCLUDED], counts[SOURCE_BACKFILL],
            ",".join(f"G{grade}" for grade in range(6) if groups[(level, grade)]), sum(row["level"] == level for row in selected),
        ])
    style_table(summary, row_index, [14, 13, 13, 15, 20, 18, 24, 13])

    coverage = wb.create_sheet("候选覆盖")
    add_title(coverage, "所选地形候选覆盖", 8)
    coverage.append([])
    coverage.append(["地形ID", "地形牌数", "难度", SOURCE_KEPT, SOURCE_EXCLUDED, SOURCE_BACKFILL, "候选合计", "已选"])
    for level in levels:
        for grade in range(6):
            candidates = groups[(level, grade)]
            counts = Counter(row["source"] for row in candidates)
            coverage.append([level, candidates[0]["tile_count"], f"G{grade}", counts[SOURCE_KEPT], counts[SOURCE_EXCLUDED], counts[SOURCE_BACKFILL], len(candidates), per_grade])
    style_table(coverage, 3, [14, 13, 10, 15, 20, 18, 13, 10])

    sources_ws = wb.create_sheet("数据源")
    add_title(sources_ws, "本次抽样数据源", 3)
    sources_ws.append([])
    sources_ws.append(["用途", "文件", "说明"])
    source_descriptions = [
        ("Strategy2 全量参数", sources[0], "提供生成参数与原始档位"),
        ("Optimal 指标", sources[1], "提供短视最优胜率、断色与失败剩余"),
        ("Optimal 保留集合", sources[2], "区分原始保留与被 Optimal 排除样本"),
        ("Optimal 补档", sources[3], "补充后续生成且通过体验约束的样本"),
    ]
    for purpose, path, description in source_descriptions:
        sources_ws.append([purpose, str(path.relative_to(ROOT)), description])
    style_table(sources_ws, 3, [22, 85, 48])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    source_counts = Counter(row["source"] for row in selected)
    return {"output": str(output), "rows": len(selected), "levels": levels, "source_counts": dict(source_counts)}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base", type=Path, default=DEFAULT_BASE)
    parser.add_argument("--metrics", type=Path, default=DEFAULT_METRICS)
    parser.add_argument("--accepted", type=Path, default=DEFAULT_ACCEPTED)
    parser.add_argument("--backfill", type=Path, default=DEFAULT_BACKFILL)
    parser.add_argument("--levels", default="100014,100088,100093")
    parser.add_argument("--per-grade", type=int, default=3)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    paths = [args.base, args.metrics, args.accepted, args.backfill]
    for path in paths:
        if not path.exists():
            raise FileNotFoundError(path)
    pool = build_pool(*paths)
    result = build_workbook(pool, [value.strip() for value in args.levels.split(",") if value.strip()], args.per_grade, args.output, paths)
    print(result)


if __name__ == "__main__":
    main()
