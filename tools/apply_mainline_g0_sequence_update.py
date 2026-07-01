#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "output"

FINAL_CSV = OUTPUT / "100003～100071_100073+_合并去少_含补缺_每档最多10.csv"
BACKFILL_CSVS = [
    OUTPUT / "主线G0补全生成_100006_100017.csv",
    OUTPUT / "主线G0补全生成_100021_100024.csv",
]
WORKBOOK = OUTPUT / "无尽关校准工具.xlsx"
REPLAY_SELECTION = OUTPUT / "replay导出" / "selection.csv"
REPORT = OUTPUT / "主线G0与Sequence更新记录.json"

GRADE_COLUMNS = [f"G{i}" for i in range(6)]
REPLAY_HEADERS = [
    "levelResId",
    "ReplayKey",
    "ReplayCode",
    "grade",
    "passrate",
    "ElementCount",
    "DifficultyScore",
    "CompletionStatus",
    "ExpectConsume",
    "LevelTags",
    "ReplayTags",
    "highWinRate",
    "MiddleWinRate",
    "LowWinRate",
]

SEQUENCE_TO_ZERO_BY_STAGE_EDITOR = {
    (21, "100021"),
    (24, "100024"),
    (78, "100070"),
}
IGNORED_STAGE_EDITORS = {
    (1, "100001"),
    (2, "100002"),
}


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def valid_row(row: dict[str, str]) -> bool:
    if not str(row.get("CompletionStatus", "Success")).startswith("Success"):
        return False
    try:
        grade = int(str(row.get("grade", "")).strip())
    except ValueError:
        return False
    return 0 <= grade <= 5


def backup(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = path.with_name(f"{path.stem}.before_mainline_g0_{stamp}{path.suffix}")
    shutil.copy2(path, dst)
    return dst


def merge_backfills() -> dict[str, object]:
    headers, rows = read_csv(FINAL_CSV)
    seen = {(r.get("levelResId", ""), r.get("ReplayCode", "")) for r in rows}
    added = []
    for path in BACKFILL_CSVS:
        _, more = read_csv(path)
        for row in more:
            key = (row.get("levelResId", ""), row.get("ReplayCode", ""))
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
            added.append(row)

    def sort_key(row: dict[str, str]) -> tuple[int, int, int]:
        try:
            level = int(row.get("levelResId", "0"))
        except ValueError:
            level = 0
        try:
            grade = int(row.get("grade", "99"))
        except ValueError:
            grade = 99
        try:
            attempt = int(row.get("attemptIndex", "0"))
        except ValueError:
            attempt = 0
        return level, grade, attempt

    rows.sort(key=sort_key)
    write_csv(FINAL_CSV, headers, rows)
    return {
        "finalCsv": str(FINAL_CSV),
        "addedRows": len(added),
        "addedByLevelGrade": {
            f"{level}_G{grade}": count
            for (level, grade), count in sorted(
                Counter((r.get("levelResId", ""), r.get("grade", "")) for r in added).items()
            )
        },
    }


def coverage_from_rows(rows: list[dict[str, str]]) -> dict[str, Counter]:
    coverage: dict[str, Counter] = defaultdict(Counter)
    for row in rows:
        if not valid_row(row):
            continue
        level = str(row.get("levelResId", "")).strip()
        if not level:
            continue
        coverage[level][int(row["grade"])] += 1
    return coverage


def grade_range(counter: Counter) -> tuple[str | None, int | None, int | None, int]:
    grades = [g for g in range(6) if counter.get(g, 0) > 0]
    total = sum(counter.get(g, 0) for g in range(6))
    if not grades:
        return None, None, None, total
    return ",".join(str(g) for g in grades), min(grades), max(grades), total


def clear_sheet_data(ws, start_row: int = 2) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def refresh_distribution_sheet(wb, coverage: dict[str, Counter]) -> None:
    ws = wb["地形G0-G5分布"]
    clear_sheet_data(ws, 2)
    for row_idx, level in enumerate(sorted(coverage, key=lambda x: int(x)), start=2):
        counter = coverage[level]
        ranges, low, high, total = grade_range(counter)
        values = [
            level,
            total,
            low,
            high,
            ranges,
            *[counter.get(g, 0) for g in range(6)],
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx).value = value


def refresh_summary_sheet(wb, coverage: dict[str, Counter], total_rows: int) -> None:
    ws = wb["难度汇总"]
    clear_sheet_data(ws, 2)
    for row_idx, grade in enumerate(range(6), start=2):
        levels = [level for level in sorted(coverage, key=lambda x: int(x)) if coverage[level].get(grade, 0) > 0]
        count = sum(coverage[level].get(grade, 0) for level in levels)
        ws.cell(row_idx, 1).value = f"G{grade}"
        ws.cell(row_idx, 2).value = len(levels)
        ws.cell(row_idx, 3).value = count
        ws.cell(row_idx, 4).value = ",".join(levels)

    if "统计说明" in wb.sheetnames:
        info = wb["统计说明"]
        values = {
            "输入行数": total_rows,
            "有效牌局行数": sum(sum(c.get(g, 0) for g in range(6)) for c in coverage.values()),
            "过滤探测行": 0,
            "过滤失败行": 0,
            "过滤无效难度行": 0,
            "地形数": len(coverage),
        }
        seen = set()
        for row in range(2, info.max_row + 1):
            key = info.cell(row, 1).value
            if key in values:
                info.cell(row, 2).value = values[key]
                seen.add(key)
        next_row = info.max_row + 1
        for key, value in values.items():
            if key in seen:
                continue
            info.cell(next_row, 1).value = key
            info.cell(next_row, 2).value = value
            next_row += 1


def refresh_mainline_coverage(wb, coverage: dict[str, Counter]) -> None:
    ws = wb["主线关难度覆盖"]
    for row in range(2, ws.max_row + 1):
        editor = str(ws.cell(row, 3).value or "").strip()
        counter = coverage.get(editor, Counter())
        ranges, low, high, total = grade_range(counter)
        ws.cell(row, 4).value = ranges
        ws.cell(row, 5).value = low
        ws.cell(row, 6).value = high
        ws.cell(row, 7).value = total
        for i, grade in enumerate(range(6), start=8):
            ws.cell(row, i).value = counter.get(grade, 0)
        ws.cell(row, 14).value = "具备" if total > 0 else "无当前覆盖数据"


def parse_required_grades(sequence: object) -> list[int]:
    text = str(sequence or "").strip()
    if not text:
        return []
    grades = []
    for part in text.replace("，", ",").split(","):
        part = part.strip()
        if not part:
            continue
        try:
            grade = int(float(part))
        except ValueError:
            continue
        if 0 <= grade <= 5:
            grades.append(grade)
    return grades


def refresh_front80(wb, coverage: dict[str, Counter]) -> list[dict[str, object]]:
    ws = wb["前80关在线胜率"]
    issues = []
    for row in range(5, ws.max_row + 1):
        editor = str(ws.cell(row, 5).value or "").strip()
        if not editor:
            continue
        stage = ws.cell(row, 1).value
        stage_key = (int(stage), editor) if isinstance(stage, int) else None
        if stage_key in SEQUENCE_TO_ZERO_BY_STAGE_EDITOR:
            ws.cell(row, 4).value = "0"
        required = parse_required_grades(ws.cell(row, 4).value)
        ws.cell(row, 6).value = ",".join(f"G{g}" for g in required)
        counter = coverage.get(editor, Counter())
        missing = [g for g in required if counter.get(g, 0) <= 0]
        if stage_key in IGNORED_STAGE_EDITORS:
            note = "无需处理"
            missing = []
        elif not counter:
            note = f"地形{editor}无当前覆盖数据"
            if missing:
                note += "；缺失" + ",".join(f"G{g}" for g in missing)
        elif missing:
            note = "缺失" + ",".join(f"G{g}" for g in missing)
        else:
            note = "具备"
        ws.cell(row, 7).value = note
        if missing or not counter:
            issues.append({
                "row": row,
                "stage": ws.cell(row, 1).value,
                "editorId": editor,
                "sequence": ws.cell(row, 4).value,
                "note": note,
            })
    return issues


def write_replay_selection(rows: list[dict[str, str]]) -> dict[str, object]:
    seen = set()
    out = []
    skipped_duplicate = 0
    for row in rows:
        if not valid_row(row):
            continue
        item = {header: row.get(header, "") for header in REPLAY_HEADERS}
        if not item["ReplayTags"]:
            item["ReplayTags"] = item["LevelTags"]
        item["LevelTags"] = ""
        key = (item["levelResId"], item["ReplayCode"])
        if key in seen:
            skipped_duplicate += 1
            continue
        seen.add(key)
        out.append(item)
    write_csv(REPLAY_SELECTION, REPLAY_HEADERS, out)
    return {
        "selectionCsv": str(REPLAY_SELECTION),
        "rows": len(out),
        "skippedDuplicateLevelReplayCode": skipped_duplicate,
        "levels": len({r["levelResId"] for r in out}),
    }


def main() -> None:
    csv_backup = backup(FINAL_CSV)
    workbook_backup = backup(WORKBOOK)

    merge_info = merge_backfills()
    headers, rows = read_csv(FINAL_CSV)
    coverage = coverage_from_rows(rows)

    wb = openpyxl.load_workbook(WORKBOOK)
    refresh_distribution_sheet(wb, coverage)
    refresh_summary_sheet(wb, coverage, len(rows))
    refresh_mainline_coverage(wb, coverage)
    front80_issues = refresh_front80(wb, coverage)
    wb.save(WORKBOOK)

    replay_info = write_replay_selection(rows)

    report = {
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
        "backups": {
            "csv": str(csv_backup),
            "workbook": str(workbook_backup),
        },
        "merge": merge_info,
        "sequenceReplacedG1ToZeroStages": [
            {"stage": stage, "editorId": editor}
            for stage, editor in sorted(SEQUENCE_TO_ZERO_BY_STAGE_EDITOR)
        ],
        "ignoredStageEditors": [
            {"stage": stage, "editorId": editor}
            for stage, editor in sorted(IGNORED_STAGE_EDITORS)
        ],
        "front80RemainingIssues": front80_issues,
        "replaySelection": replay_info,
        "configJson": {
            "path": str(OUTPUT / "关卡配置B_填充无尽配置.json"),
            "changed": False,
            "reason": "本次补的是主线关 editorId，不在当前无尽 LevelPool；无尽配置无需改动。",
        },
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
