#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "output"
LEVELS_DIR = ROOT.parent / "TileMatchShell" / "Tools" / "Config" / "Json" / "Levels"


def normalize_sequence(value: Any) -> str:
    text = str(value or "").strip().replace("，", ",")
    if not text:
        return ""
    parts: list[str] = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            number = int(float(part))
        except ValueError:
            parts.append(part)
            continue
        parts.append(str(number))
    return ",".join(parts)


def read_table_rows(workbook_path: Path, sheet_name: str, header_row: int) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {str(header): ws.cell(r, c).value for c, header in enumerate(headers, start=1) if header}
        if any(value is not None for value in row.values()):
            rows.append(row)
    return rows


def read_front80_sequences(workbook_path: Path) -> dict[int, dict[str, str]]:
    rows = read_table_rows(workbook_path, "前80关在线胜率", 4)
    result: dict[int, dict[str, str]] = {}
    for row in rows:
        level = row.get("关卡ID")
        editor = row.get("EditorID")
        sequence = row.get("近似模拟GradeSequence")
        if level is None or editor is None:
            continue
        try:
            level_int = int(level)
        except (TypeError, ValueError):
            continue
        result[level_int] = {
            "LevelResID": str(editor).strip(),
            "GradeSequence": normalize_sequence(sequence),
            "Remark": str(row.get("备注") or "").strip(),
        }
    return result


def read_zone_sequences(workbook_path: Path) -> list[list[str]]:
    rows = read_table_rows(workbook_path, "无尽难度集", 4)
    zones: list[list[str]] = []
    for row in rows:
        group = str(row.get("组") or "").strip()
        if not group:
            break
        if len(group) != 1 or not ("A" <= group.upper() <= "Z"):
            continue
        seqs = [
            normalize_sequence(row.get("第1关序列")),
            normalize_sequence(row.get("第2关序列")),
            normalize_sequence(row.get("第3关序列")),
            normalize_sequence(row.get("第4关序列")),
            normalize_sequence(row.get("第5关序列")),
        ]
        if any(seqs):
            zones.append(seqs)
    return zones


def parse_grade_list(value: Any) -> list[int]:
    text = str(value or "").strip().replace("，", ",")
    grades: list[int] = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            grade = int(float(part))
        except ValueError:
            continue
        if grade not in grades:
            grades.append(grade)
    return grades


def tile_count(level_res_id: str) -> int:
    data = json.loads((LEVELS_DIR / f"{level_res_id}.json").read_text(encoding="utf-8"))
    return sum(len(layer.get("tiles", [])) for layer in data.get("layers", []))


def keep_grade(grade: int, tiles: int) -> bool:
    if grade in (1, 2):
        return tiles <= 90
    if grade == 3:
        return 66 <= tiles <= 90
    if grade in (4, 5):
        return tiles >= 72
    return False


def read_filtered_level_pool(workbook_path: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    rows = read_table_rows(workbook_path, "无尽关难度覆盖", 1)
    pool: list[dict[str, Any]] = []
    details: dict[str, dict[str, Any]] = {}
    for row in rows:
        editor = str(row.get("EditorID") or "").strip()
        if not editor:
            continue
        tiles = tile_count(editor)
        original = parse_grade_list(row.get("可用难度"))
        filtered = [grade for grade in original if keep_grade(grade, tiles)]
        details[editor] = {
            "tileCount": tiles,
            "originalGrades": original,
            "filteredGrades": filtered,
        }
        if filtered:
            pool.append({
                "LevelResID": editor,
                "GradeRange": filtered,
            })
    return pool, details


def build_stage_overrides(
    baseline: dict[int, dict[str, str]],
    target: dict[int, dict[str, str]],
    valid_length: int,
) -> list[dict[str, Any]]:
    overrides: list[dict[str, Any]] = []
    for level in range(1, valid_length + 1):
        base = baseline.get(level)
        next_value = target.get(level)
        if not next_value:
            continue
        if next_value.get("Remark") == "无需处理":
            continue
        base_sequence = base["GradeSequence"] if base else ""
        target_sequence = next_value["GradeSequence"]
        if not target_sequence or target_sequence == base_sequence:
            continue
        overrides.append({
            "level": level,
            "LevelResID": next_value["LevelResID"],
            "GradeSequence": target_sequence,
        })
    return overrides


def update_zones(config: dict[str, Any], zone_sequences: list[list[str]]) -> None:
    zones = config.get("Zones")
    if not isinstance(zones, list):
        config["Zones"] = []
        zones = config["Zones"]
    for idx, sequences in enumerate(zone_sequences):
        if idx < len(zones) and isinstance(zones[idx], dict):
            zone = zones[idx]
        else:
            zone = {
                "zoneId": idx + 1,
                "name": f"难度曲线{chr(ord('A') + idx)}",
                "gradeSequences": [],
            }
            zones.append(zone)
        zone["gradeSequences"] = sequences
    del zones[len(zone_sequences):]


def update_workbook_copy(workbook_path: Path, filtered: dict[str, dict[str, Any]]) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["无尽关难度覆盖"]
    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    for row in range(2, ws.max_row + 1):
        editor = str(ws.cell(row, headers["EditorID"]).value or "").strip()
        info = filtered.get(editor)
        if not info:
            continue
        grades = info["filteredGrades"]
        ws.cell(row, headers["可用难度"]).value = ",".join(str(g) for g in grades) if grades else None
        ws.cell(row, headers["最低难度"]).value = min(grades) if grades else None
        ws.cell(row, headers["最高难度"]).value = max(grades) if grades else None
        total = 0
        for grade in range(0, 6):
            col = headers.get(f"G{grade}")
            if not col:
                continue
            if grade in grades:
                value = ws.cell(row, col).value or 0
            else:
                value = 0
            ws.cell(row, col).value = value
            total += int(value or 0)
        ws.cell(row, headers["总牌局数"]).value = total
        ws.cell(row, headers["备注"]).value = "具备" if grades else "无tile规则可用难度"

    try:
        from build_calibration_variant import refresh_simulation
        refresh_simulation(wb)
    except Exception:
        # Keep the workbook usable even if simulation refresh is unavailable.
        pass
    wb.save(workbook_path)


def safe_name(text: str) -> str:
    return re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", text).strip("_")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export a tile-count experience strategy package.")
    parser.add_argument("--strategy-name", default="tile数体验校准")
    parser.add_argument("--baseline-workbook", type=Path, default=OUTPUT / "无尽关校准工具.xlsx")
    parser.add_argument("--target-workbook", type=Path, default=OUTPUT / "无尽关校准工具_无补缺.xlsx")
    parser.add_argument("--config", type=Path, default=OUTPUT / "关卡配置B_0626增难版.json")
    parser.add_argument("--out-root", type=Path, default=OUTPUT / "strategy_runs")
    parser.add_argument("--stage-valid-length", type=int, default=80)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = args.out_root / f"{stamp}_{safe_name(args.strategy_name)}"
    output_dir = run_dir / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    level_pool, filter_details = read_filtered_level_pool(args.target_workbook)
    baseline = read_front80_sequences(args.baseline_workbook)
    target = read_front80_sequences(args.target_workbook)
    zone_sequences = read_zone_sequences(args.target_workbook)
    overrides = build_stage_overrides(baseline, target, args.stage_valid_length)

    config = json.loads(args.config.read_text(encoding="utf-8"))
    updated = deepcopy(config)
    updated["LevelPool"] = level_pool
    stage_override = updated.setdefault("StageOverride", {})
    stage_override["StageValidLength"] = str(args.stage_valid_length)
    stage_override["Stages"] = overrides
    update_zones(updated, zone_sequences)

    config_out = output_dir / "关卡配置B_tile数体验校准.json"
    config_out.write_text(json.dumps(updated, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    workbook_out = output_dir / "tile数体验校准.xlsx"
    shutil.copy2(args.target_workbook, workbook_out)
    update_workbook_copy(workbook_out, filter_details)

    manifest = {
        "strategyName": args.strategy_name,
        "createdAt": datetime.now().isoformat(timespec="seconds"),
        "rules": {
            "G1": "tile <= 90",
            "G2": "tile <= 90",
            "G3": "66 <= tile <= 90",
            "G4": "tile >= 72",
            "G5": "tile >= 72",
            "G0": "not used",
        },
        "inputs": {
            "sourceConfig": str(args.config),
            "baselineWorkbook": str(args.baseline_workbook),
            "targetWorkbook": str(args.target_workbook),
        },
        "outputs": {
            "config": str(config_out),
            "workbook": str(workbook_out),
        },
        "summary": {
            "levelPoolCount": len(level_pool),
            "stageValidLength": args.stage_valid_length,
            "stageOverrideCount": len(overrides),
            "zoneCount": len(zone_sequences),
            "gradeCoverage": {
                str(grade): sum(1 for item in level_pool if grade in item["GradeRange"])
                for grade in range(1, 6)
            },
        },
    }
    manifest_out = run_dir / "manifest.json"
    manifest_out.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    index_out = args.out_root / "index.jsonl"
    with index_out.open("a", encoding="utf-8") as f:
        f.write(json.dumps({
            "createdAt": manifest["createdAt"],
            "strategyName": args.strategy_name,
            "runDir": str(run_dir),
            "config": str(config_out),
            "workbook": str(workbook_out),
            "summary": manifest["summary"],
        }, ensure_ascii=False) + "\n")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
