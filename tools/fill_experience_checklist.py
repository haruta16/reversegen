#!/usr/bin/env python3
"""Fill the existing experience checklist while preserving its layout."""

from __future__ import annotations

import argparse
import os
import shutil
from collections import Counter, defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from build_experience_sample_workbook import (
    DEFAULT_ACCEPTED,
    DEFAULT_BACKFILL,
    DEFAULT_BASE,
    DEFAULT_METRICS,
    ROOT,
    SOURCE_BACKFILL,
    SOURCE_EXCLUDED,
    SOURCE_KEPT,
    build_pool,
    difference_label,
    read_csv,
    select_diverse,
)


DEFAULT_CHECKLIST = ROOT / "output/体验清单.xlsx"
DEFAULT_FINAL_SELECTION = ROOT / "output/strategy_runs/20260630_至少8局当前校准/03_replay/selection.csv"


def copy_cell_style(source, target) -> None:
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.protection = copy(source.protection)
    target.alignment = copy(source.alignment)


def expected_headers(ws) -> list[str]:
    return [ws.cell(1, 1).value, ws.cell(1, 2).value, ws.cell(1, 3).value] + [ws.cell(2, column).value for column in range(4, 11)]


def fill_checklist(template: Path, levels: list[str], per_grade: int, final_selection: Path) -> dict[str, object]:
    pool = build_pool(DEFAULT_BASE, DEFAULT_METRICS, DEFAULT_ACCEPTED, DEFAULT_BACKFILL)
    final_codes = {row["ReplayCode"] for row in read_csv(final_selection)}
    groups = defaultdict(list)
    for row in pool:
        if row["level"] in levels and 0 <= row["grade_num"] <= 5:
            groups[(row["level"], row["grade_num"])].append(row)

    missing = [
        (level, grade, len(groups[(level, grade)]))
        for level in levels
        for grade in range(6)
        if len(groups[(level, grade)]) < per_grade
    ]
    if missing:
        raise ValueError(f"候选不足: {missing}")

    selected = []
    for level in levels:
        for grade in range(6):
            candidates = groups[(level, grade)]
            for row in select_diverse(candidates, per_grade):
                picked = dict(row)
                picked["difference"] = difference_label(row, candidates)
                selected.append(picked)

    workbook = load_workbook(template)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = expected_headers(worksheet)
    required = [
        "关卡号", "Replaycode", "实际难度分档", "花色数", "实际花色系数", "花色分布参数",
        "债务持续参数", "短视最优策略胜率", "胜局思考量（断色数）", "失败剩余率",
    ]
    if headers != required:
        raise ValueError(f"体验清单表头不匹配: {headers}")

    for merged in list(worksheet.merged_cells.ranges):
        if str(merged).startswith("A") and merged.min_row >= 3:
            worksheet.unmerge_cells(str(merged))

    first_data_row = 3
    total_rows = len(selected)
    last_data_row = first_data_row + total_rows - 1
    template_rows = list(range(3, 13))
    for target_row in range(first_data_row, last_data_row + 1):
        source_row = template_rows[(target_row - first_data_row) % len(template_rows)]
        for column in range(1, 11):
            copy_cell_style(worksheet.cell(source_row, column), worksheet.cell(target_row, column))
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
        for column in range(1, 11):
            cell = worksheet.cell(target_row, column)
            cell.value = None
            cell.comment = None

    for row in range(last_data_row + 1, worksheet.max_row + 1):
        for column in range(1, 11):
            worksheet.cell(row, column).value = None
            worksheet.cell(row, column).comment = None

    level_fills = {
        levels[0]: PatternFill("solid", fgColor="EAF2F8"),
        levels[1]: PatternFill("solid", fgColor="EAF7EE"),
        levels[2]: PatternFill("solid", fgColor="FFF3E6"),
    }
    grade_colors = {0: "1A7F37", 1: "39814B", 2: "9A6700", 3: "B65D20", 4: "C2413B", 5: "A40E26"}
    source_counts = Counter()
    final_selected_count = 0
    cursor = first_data_row
    for level in levels:
        level_start = cursor
        level_rows = [row for row in selected if row["level"] == level]
        for row in level_rows:
            is_final_selected = row["ReplayCode"] in final_codes
            values = [
                None,
                row["ReplayCode"],
                f"✓ G{row['grade_num']}" if is_final_selected else f"G{row['grade_num']}",
                row["color_count"],
                row["color_ratio"],
                row["spread"],
                row["debt"],
                row["optimal_win"],
                row["starvation_win"],
                row["remaining_loss"],
            ]
            for column, value in enumerate(values, 1):
                worksheet.cell(cursor, column, value)
            worksheet.cell(cursor, 2).comment = Comment(
                f"样本来源：{row['source']}\n差异特征：{row['difference']}\n"
                f"正式资源：{'✓ 已入选' if is_final_selected else '未入选（体验对照）'}\n"
                f"候选范围：地形 {level} / G{row['grade_num']}",
                "Codex",
            )
            worksheet.cell(cursor, 3).comment = Comment(
                "✓ 表示 ReplayCode 存在于当前正式资源包 selection.csv。" if is_final_selected else "该 ReplayCode 不在当前正式资源包中，仅用于体验对照。",
                "Codex",
            )
            if is_final_selected:
                worksheet.cell(cursor, 3).fill = PatternFill("solid", fgColor="2E7D32")
                worksheet.cell(cursor, 3).font = Font(name="PingFang SC", size=worksheet.cell(cursor, 3).font.sz or 11, bold=True, color="FFFFFF")
                final_selected_count += 1
            else:
                worksheet.cell(cursor, 3).font = Font(
                    name="PingFang SC", size=worksheet.cell(cursor, 3).font.sz or 11, bold=True, color=grade_colors[row["grade_num"]]
                )
            worksheet.cell(cursor, 5).number_format = "0.000"
            worksheet.cell(cursor, 6).number_format = "0.000"
            worksheet.cell(cursor, 7).number_format = "0.000"
            worksheet.cell(cursor, 8).number_format = "0.0%"
            worksheet.cell(cursor, 9).number_format = "0.00"
            worksheet.cell(cursor, 10).number_format = "0.0%"
            source_counts[row["source"]] += 1
            cursor += 1
        level_end = cursor - 1
        worksheet.merge_cells(start_row=level_start, start_column=1, end_row=level_end, end_column=1)
        level_cell = worksheet.cell(level_start, 1, level)
        level_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        level_cell.font = Font(name="PingFang SC", bold=True, size=12)
        level_cell.fill = level_fills[level]

    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"B2:J{last_data_row}"
    worksheet.cell(1, 4).comment = Comment(
        "ReplayCode 批注中记录样本来源和差异特征。\n"
        "失败剩余率仅在 Optimal 有失败局时填写；全胜牌局留空。",
        "Codex",
    )
    worksheet.cell(1, 3).comment = Comment(
        "绿色的“✓ Gx”表示该 ReplayCode 存在于当前正式资源包 selection.csv；\n"
        "没有 ✓ 的牌局只作为体验对照，不在正式资源中。",
        "Codex",
    )

    backup = template.with_name(f"{template.stem}_填写前备份_{datetime.now():%Y%m%d_%H%M%S}{template.suffix}")
    shutil.copy2(template, backup)
    temporary = template.with_name(f".{template.name}.tmp.xlsx")
    workbook.save(temporary)
    os.replace(temporary, template)
    return {
        "file": str(template),
        "backup": str(backup),
        "rows": total_rows,
        "levels": levels,
        "source_counts": dict(source_counts),
        "final_selected": final_selected_count,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", type=Path, default=DEFAULT_CHECKLIST)
    parser.add_argument("--levels", default="100014,100088,100093")
    parser.add_argument("--per-grade", type=int, default=3)
    parser.add_argument("--final-selection", type=Path, default=DEFAULT_FINAL_SELECTION)
    args = parser.parse_args()
    levels = [value.strip() for value in args.levels.split(",") if value.strip()]
    if len(levels) != 3:
        raise ValueError("体验清单当前版式要求正好三个地形")
    print(fill_checklist(args.file, levels, args.per_grade, args.final_selection))


if __name__ == "__main__":
    main()
