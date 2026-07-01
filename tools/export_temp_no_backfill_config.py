#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "output"


def normalize_sequence(value: Any) -> str:
    text = str(value or "").strip().replace("，", ",")
    if not text:
        return ""
    parts: list[str] = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            number = int(float(part))
        except ValueError:
            parts.append(part)
            continue
        parts.append(str(number))
    return ",".join(parts)


def read_table_rows(workbook_path: Path, sheet_name: str, header_row: int) -> list[dict[str, Any]]:
    # Artifact Tool exports may omit the optional worksheet dimension hint.
    # Normal mode derives dimensions from cells; read-only mode can return None.
    wb = load_workbook(workbook_path, data_only=True, read_only=False)
    ws = wb[sheet_name]
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {str(header): ws.cell(r, c).value for c, header in enumerate(headers, start=1) if header}
        if any(value is not None for value in row.values()):
            rows.append(row)
    return rows


def read_front80_sequences(workbook_path: Path) -> dict[int, dict[str, str]]:
    rows = read_table_rows(workbook_path, "前80关在线胜率", 4)
    result: dict[int, dict[str, str]] = {}
    for row in rows:
        level = row.get("关卡ID")
        editor = row.get("EditorID")
        sequence = row.get("近似模拟GradeSequence")
        if level is None or editor is None:
            continue
        try:
            level_int = int(level)
        except (TypeError, ValueError):
            continue
        result[level_int] = {
            "LevelResID": str(editor).strip(),
            "GradeSequence": normalize_sequence(sequence),
            "Remark": str(row.get("备注") or "").strip(),
        }
    return result


def read_zone_sequences(workbook_path: Path) -> list[list[str]]:
    rows = read_table_rows(workbook_path, "无尽难度集", 4)
    zones: list[list[str]] = []
    for row in rows:
        group = str(row.get("组") or "").strip()
        if not group:
            break
        # Only the top editable zone table uses groups A-F. The same sheet also
        # contains formula breakdown rows below it; do not export those as zones.
        if len(group) != 1 or not ("A" <= group.upper() <= "Z"):
            continue
        seqs = [
            normalize_sequence(row.get("第1关序列")),
            normalize_sequence(row.get("第2关序列")),
            normalize_sequence(row.get("第3关序列")),
            normalize_sequence(row.get("第4关序列")),
            normalize_sequence(row.get("第5关序列")),
        ]
        if any(seqs):
            zones.append(seqs)
    return zones


def parse_grade_list(value: Any) -> list[int]:
    text = str(value or "").strip().replace("，", ",")
    grades: list[int] = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            grade = int(float(part))
        except ValueError:
            continue
        if grade not in grades:
            grades.append(grade)
    return grades


def read_level_pool(workbook_path: Path) -> list[dict[str, Any]]:
    rows = read_table_rows(workbook_path, "无尽关难度覆盖", 1)
    pool: list[dict[str, Any]] = []
    for row in rows:
        editor = str(row.get("EditorID") or "").strip()
        if not editor:
            continue
        grades = parse_grade_list(row.get("可用难度"))
        if not grades:
            continue
        pool.append({
            "LevelResID": editor,
            "GradeRange": grades,
        })
    return pool


def build_stage_overrides(
    baseline: dict[int, dict[str, str]],
    target: dict[int, dict[str, str]],
    valid_length: int,
) -> list[dict[str, Any]]:
    overrides: list[dict[str, Any]] = []
    for level in range(1, valid_length + 1):
        base = baseline.get(level)
        next_value = target.get(level)
        if not next_value:
            continue
        if next_value.get("Remark") == "无需处理":
            continue
        base_sequence = base["GradeSequence"] if base else ""
        target_sequence = next_value["GradeSequence"]
        if not target_sequence or target_sequence == base_sequence:
            continue
        overrides.append({
            "level": level,
            "LevelResID": next_value["LevelResID"],
            "GradeSequence": target_sequence,
        })
    return overrides


def update_zones(config: dict[str, Any], zone_sequences: list[list[str]]) -> None:
    zones = config.get("Zones")
    if not isinstance(zones, list):
        config["Zones"] = []
        zones = config["Zones"]

    for idx, sequences in enumerate(zone_sequences):
        if idx < len(zones) and isinstance(zones[idx], dict):
            zone = zones[idx]
        else:
            zone = {
                "zoneId": idx + 1,
                "name": f"难度曲线{chr(ord('A') + idx)}",
                "gradeSequences": [],
            }
            zones.append(zone)
        zone["gradeSequences"] = sequences

    del zones[len(zone_sequences):]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Temporary export: compare no-backfill mainline GradeSequence against filled baseline and update config overrides/zones.",
    )
    parser.add_argument("--baseline-workbook", type=Path, default=OUTPUT / "无尽关校准工具.xlsx")
    parser.add_argument("--target-workbook", type=Path, default=OUTPUT / "无尽关校准工具_无补缺.xlsx")
    parser.add_argument("--config", type=Path, default=OUTPUT / "关卡配置B_填充无尽配置.json")
    parser.add_argument("--output", type=Path, default=OUTPUT / "关卡配置B_无补缺底板_临时.json")
    parser.add_argument("--report", type=Path, default=None)
    parser.add_argument("--stage-valid-length", type=int, default=80)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    baseline = read_front80_sequences(args.baseline_workbook)
    target = read_front80_sequences(args.target_workbook)
    level_pool = read_level_pool(args.target_workbook)
    zone_sequences = read_zone_sequences(args.target_workbook)
    overrides = build_stage_overrides(baseline, target, args.stage_valid_length)

    config = json.loads(args.config.read_text(encoding="utf-8"))
    updated = deepcopy(config)
    updated["LevelPool"] = level_pool
    stage_override = updated.setdefault("StageOverride", {})
    stage_override["StageValidLength"] = str(args.stage_valid_length)
    stage_override["Stages"] = overrides
    update_zones(updated, zone_sequences)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(updated, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    summary = {
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
        "baselineWorkbook": str(args.baseline_workbook),
        "targetWorkbook": str(args.target_workbook),
        "sourceConfig": str(args.config),
        "outputConfig": str(args.output),
        "levelPoolCount": len(level_pool),
        "stageValidLength": args.stage_valid_length,
        "stageOverrideCount": len(overrides),
        "zoneCount": len(zone_sequences),
        "zones": zone_sequences,
    }
    if args.report:
        args.report.write_text(json.dumps({
            **summary,
            "stageOverrides": overrides,
        }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
